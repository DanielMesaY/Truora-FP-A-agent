"""
extract.py — Paso de EXTRACCIÓN del pipeline.
Lee un estado de resultados mensual (.xlsx) y devuelve un dict normalizado de cifras.

Extracción por ETIQUETA (no por celda fija) para tolerar pequeños cambios de layout.
En n8n: nodo Code (Python) o un microservicio que reciba el archivo de Google Drive.
"""
import unicodedata
from openpyxl import load_workbook

# Mapa: etiqueta normalizada del P&G -> clave canónica
LABEL_MAP = {
    "ingresos operacionales": "ingresos",
    "costo de ventas": "costo_ventas",
    "utilidad bruta": "utilidad_bruta",
    "margen bruto": "margen_bruto",
    "gastos de administracion": "gastos_admin",
    "gastos de ventas": "gastos_ventas",
    "gastos de mercadeo": "gastos_mercadeo",
    "ebitda": "ebitda",
    "margen ebitda": "margen_ebitda",
    "depreciacion y amortizacion": "depreciacion_amortizacion",
    "utilidad operacional ebit": "ebit",
    "margen operacional": "margen_operacional",
    "ingresos financieros": "ingresos_fin",
    "gastos financieros": "gastos_fin",
    "utilidad antes de impuestos": "uai",
    "impuesto de renta": "impuesto_renta",
    "utilidad neta": "utilidad_neta",
    "margen neto": "margen_neto",
}

def _norm(s: str) -> str:
    """minúsculas, sin acentos, sin signos (-, %, +, (), espacios colapsados)."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = s.lower()
    for ch in "-%+().":
        s = s.replace(ch, " ")
    return " ".join(s.split())

def extract_income_statement(path: str) -> dict:
    """Devuelve un dict con las 18 líneas canónicas del P&G del mes."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    out = {}
    for row in ws.iter_rows(min_col=1, max_col=2):
        label_cell, value_cell = row[0], row[1]
        key = LABEL_MAP.get(_norm(label_cell.value))
        if key and isinstance(value_cell.value, (int, float)):
            out[key] = float(value_cell.value)
    missing = set(LABEL_MAP.values()) - set(out.keys())
    if missing:
        # No inventamos cifras: reportamos lo que falta (principio del prompt).
        raise ValueError(f"Faltan líneas en {path}: {sorted(missing)}")
    return out

if __name__ == "__main__":
    import sys, json
    print(json.dumps(extract_income_statement(sys.argv[1]), indent=2, ensure_ascii=False))
